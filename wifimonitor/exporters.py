import subprocess
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook


class HashcatExporter:
    def __init__(self, tool_path: str = "hcxpcapngtool") -> None:
        self.tool_path = tool_path

    def export(self, capture_path: Path, output_path: Path) -> None:
        command = [self.tool_path, "-o", str(output_path), str(capture_path)]
        try:
            result = subprocess.run(command, capture_output=True, text=True, check=False)
        except FileNotFoundError as exc:
            raise FileNotFoundError(
                f"Утилита '{self.tool_path}' не найдена. Установите hcxpcapngtool или укажите путь к ней."
            ) from exc
        if result.returncode != 0:
            error = result.stderr.strip() or result.stdout.strip() or "Failed to export handshake"
            raise RuntimeError(error)


class ExcelExporter:
    def export(
        self,
        output_path: Path,
        access_points: Sequence[dict],
        stations: Sequence[dict],
        handshakes: Sequence[dict],
    ) -> None:
        workbook = Workbook()
        ws_ap = workbook.active
        ws_ap.title = "AccessPoints"
        ws_ap.append(["BSSID", "ESSID", "Channel", "Encryption", "Signal", "Last Seen"])
        for ap in access_points:
            ws_ap.append([
                ap.get("bssid"),
                ap.get("essid"),
                ap.get("channel"),
                ap.get("encryption"),
                ap.get("signal"),
                ap.get("last_seen"),
            ])

        ws_stations = workbook.create_sheet("Stations")
        ws_stations.append(["MAC", "Associated BSSID", "Signal", "Last Seen"])
        for station in stations:
            ws_stations.append([
                station.get("mac"),
                station.get("associated_bssid"),
                station.get("signal"),
                station.get("last_seen"),
            ])

        ws_handshakes = workbook.create_sheet("Handshakes")
        ws_handshakes.append(["BSSID", "Station", "Capture Path", "Created At"])
        for handshake in handshakes:
            ws_handshakes.append([
                handshake.get("bssid"),
                handshake.get("station_mac"),
                handshake.get("capture_path"),
                handshake.get("created_at"),
            ])

        workbook.save(output_path)
